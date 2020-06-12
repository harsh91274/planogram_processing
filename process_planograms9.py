import pandas as pd
import numpy as np
import os
import xlsxwriter 
import openpyxl as xl
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore');

#instructions
#no empty lines on top of planogram pages
#pages must be named as per convention

#filename=('April 2020 Planogram of Home Depot Fisher.xlsx');
filename= input("Enter File Name (add extension):");
outfile=pd.ExcelWriter("processed_"+filename,engine='xlsxwriter'); 

path1 = os.path.join(os.getcwd(), filename);
path2 = os.path.join(os.getcwd(), outfile);
    
#ingest all the data
wb=load_workbook(filename,read_only=True); 
fields0=["Brand / Appliance","Model #", "MSRP", "Occurrences", "Planogram_row", "Planogram_column"];
fields1=["Sale Price", "Assigned Value"]; 
pre=filename.split(" ")[0]+" "+filename.split(" ")[1];
fields2=[pre+" "+post for post in fields1];
fields=fields0+fields2;

default=[str("none"),str("none"),0,0,str("none"),str("none"), 0, 0];
df_def=pd.DataFrame([default],columns=fields);
#%%
wb1_c=0;
if 'Planogram' in wb.sheetnames:
    df=pd.read_excel(filename,'Planogram', header=None, skiprows=None); 
    wb1_c=1;
    print('Planogram found');
else:
    print('Planogram does not exist');

wb2_c=0;
if 'Washers-Dryers' in wb.sheetnames:
    df_washd=pd.read_excel(filename,'Washers-Dryers'); 
    wb2_c=1;
else:
    print('Washers-Dryers information does not exist');
    df_washd=pd.DataFrame(columns=fields);  
    
wb3_c=0;
if 'Dishwashers' in wb.sheetnames:
    df_dishw=pd.read_excel(filename,'Dishwashers');  
    wb3_c=1;
else:
    print('Dishwashers information does not exist');
    df_dishw=pd.DataFrame(columns=fields);

wb4_c=0;
if 'Refrigerators' in wb.sheetnames:
    df_ref=pd.read_excel(filename,'Refrigerators');   
    wb4_c=1;
else:
    print('Refrigerators information does not exist');
    df_ref=pd.DataFrame(columns=fields);

wb5_c=0;
if 'Cooking Products' in wb.sheetnames:
    df_cookp=pd.read_excel(filename,'Cooking Products');   
    wb5_c=1;
else:
    print('Cooking Products information does not exist');
    df_cookp=pd.DataFrame(columns=fields);

#ploc_* have two columns: c1 is number of occurences, c2 is planogram location(s)
[plan_rows, plan_cols]=df.shape;
#%%
#washer dryer planogram location
if wb2_c==1: 
    n_washd=df_washd.shape[0]; freq_washd=np.empty((n_washd,1)); pos_washd=["" for x in range (n_washd)]; row_washd=["" for x in range (n_washd)];
    print('WASHER-DRYERS');
    for i in range (n_washd):
        str_model=str(df_washd.loc[i,'Model #']); 
        count=0;
        locs=[];
        rowsa=[];
            
        for row in range (df.shape[0]):
            for col in range(df.shape[1]):
                if df.astype(str).iloc[row,col].count(str_model)>0 and str_model!='nan':
                    count=count+1;
                    locs.append(str(xlsxwriter.utility.xl_col_to_name(col)));
                    rowsa.append(str(row+1));
                
        freq_washd[i]=count;
        pos_washd[i]=','.join(locs); 
        row_washd[i]=','.join(rowsa);
        print("Count for ", str_model, ": ", count, ", planogram row: ", str(row_washd[i]),", column: ",  pos_washd[i]);    
    
    df_washd['Occurrences']=freq_washd;
    df_washd['Planogram_row']=row_washd;    
    df_washd['Planogram_column']=pos_washd;
#else:
    #df_washd=df_def;
    
df_washd.to_excel(outfile,'Washers-Dryers', index=False);
#outfile.save();
#%%
#dishwasher planogram location
if wb3_c==1:
    n_dishw=df_dishw.shape[0]; freq_dishw=np.empty((n_dishw,1)); pos_dishw=["" for x in range (n_dishw)];row_dishw=["" for x in range (n_dishw)];
    print('DISHWASHERS');
    for i in range (n_dishw):
        str_model=str(df_dishw.loc[i,'Model #']); 
        count=0;
        locs=[];
        rowsa=[];
        
        for row in range (df.shape[0]):
            for col in range(df.shape[1]):
                if df.astype(str).iloc[row,col].count(str_model)>0 and str_model!='nan':
                    count=count+1;
                    locs.append(str(xlsxwriter.utility.xl_col_to_name(col)));
                    rowsa.append(str(row+1));
                
        freq_dishw[i]=count;
        pos_dishw[i]=','.join(locs);    
        #print("Count for ", str_model, ": ", count, ", planogram position: ",pos_dishw[i]);
        row_dishw[i]=','.join(rowsa);
        print("Count for ", str_model, ": ", count, ", planogram row: ", str(row_dishw[i]),", column: ",  pos_dishw[i]);    
            
    df_dishw['Occurrences']=freq_dishw;
    df_dishw['Planogram_row']=row_dishw; 
    df_dishw['Planogram_column']=pos_dishw;
#else:
   # df_dishw=df_def;
    
df_dishw.to_excel(outfile,'Dishwashers', index=False);
#outfile.save();
#%%
#refrigerator planogram location
if wb4_c==1:
    n_ref=df_ref.shape[0]; freq_ref=np.empty((n_ref,1)); pos_ref=["" for x in range (n_ref)]; row_ref=["" for x in range (n_ref)];    
    print('REFRIGERATORS');
    for i in range (n_ref):
        str_model=str(df_ref.loc[i,'Model #']); 
        count=0;
        locs=[];
        rowsa=[];
            
        for row in range (df.shape[0]):
            for col in range(df.shape[1]):
                if df.astype(str).iloc[row,col].count(str_model)>0 and str_model!='nan':
                    count=count+1;
                    locs.append(str(xlsxwriter.utility.xl_col_to_name(col)));
                    rowsa.append(str(row+1));
                
        freq_ref[i]=count;
        pos_ref[i]=','.join(locs);    
        #print("Count for ", str_model, ": ", count, ", planogram position: ",pos_ref[i]);    
        row_ref[i]=','.join(rowsa);
        print("Count for ", str_model, ": ", count, ", planogram row: ", str(row_ref[i]),", column: ",  pos_ref[i]);    
        
    df_ref['Occurrences']=freq_ref;
    df_ref['Planogram_row']=row_ref;
    df_ref['Planogram_column']=pos_ref;
#else:
   # df_ref=df_def;
    
df_ref.to_excel(outfile,'Refrigerators', index=False);
#outfile.save();
#%%
#cooking products planogram location
if wb5_c==1:
    n_cookp=df_cookp.shape[0]; freq_cookp=np.empty((n_cookp,1)); pos_cookp=["" for x in range (n_cookp)];row_cookp=["" for x in range (n_cookp)];
    print('Cooking Products');
    for i in range (n_cookp):
        str_model=str(df_cookp.loc[i,'Model #']); 
        count=0;
        locs=[];
        rowsa=[];
            
        for row in range (df.shape[0]):
            for col in range(df.shape[1]):
                if df.astype(str).iloc[row,col].count(str_model)>0 and str_model!='nan':
                    count=count+1;
                    locs.append(str(xlsxwriter.utility.xl_col_to_name(col)));
                    rowsa.append(str(row+1));
                    
        freq_cookp[i]=count;
        pos_cookp[i]=','.join(locs);    
        #print("Count for ", str_model, ": ", count, ", planogram position: ",pos_cookp[i]);    
        row_cookp[i]=','.join(rowsa);
        print("Count for ", str_model, ": ", count, ", planogram row: ", str(row_cookp[i]),", column: ",  pos_cookp[i]);    
         
    df_cookp['Occurrences']=freq_cookp;
    df_cookp['Planogram_row']=row_cookp;  
    df_cookp['Planogram_column']=pos_cookp;
#else:
  #  df_cookp=df_def;
    
df_cookp.to_excel(outfile,'Cooking Products', index=False);
outfile.save();
#%%
#copying the planogram
if wb1_c==1:
    wb1 = xl.load_workbook(filename=path1)
    ws1 = wb1.worksheets[0]
    wb2 = xl.load_workbook(filename=path2)
    ws2 = wb2.create_sheet("Planogram")
        
    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value
    
    wb2.save(path2)